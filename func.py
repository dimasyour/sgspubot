#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pandas as pd
import vk_api
from openpyxl import load_workbook
from vk_api.longpoll import *
from dbworker import *
from datetime import datetime as dt
import datetime

print(colorama.Fore.LIGHTBLUE_EX + 'Парсинг новостей - успешно!')


def parse_table_spec():
    url = "https://www.samgups.ru/education/abiturientam/pk-2020-vo/samara/spetsialnosti_i_napravleniya_podgotovki/index.php"
    table = pd.read_html(url)[0]
    table.to_excel("src/ochnik.xlsx")
    table = pd.read_html(url)[1]
    table.to_excel("src/zaochnik.xlsx")


def parse_table_sgspu():
    url = "http://www.pgsga.ru/abitur/bachelor/#abitur_spiski-postupayushchih"
    table = pd.read_html(url)[0]
    table.to_excel("src/sgspu.xlsx")
    table = pd.read_html(url)[1]
    table.to_excel("src/sgspu_1.xlsx")


def view_spec(file):
    wb1 = load_workbook(file)
    sheet_ranges = wb1['Sheet1']
    column_b = sheet_ranges['B']
    column_c = sheet_ranges['C']
    column_d = sheet_ranges['D']

    specList = []
    for i in range(5, len(column_c)):
        if column_b[i].value != 'БАКАЛАВРИАТ (срок обучения - 4 года)':
            f = ' ' + column_d[i].value.replace(',', '\n')
            specList.append('❗' + column_b[i].value + '\n✅ ' + column_c[i].value + ': \n' + f + '\n\n')
        elif column_b[i].value != '':
            break
    specString = '\n'.join(specList)
    return specString


# парситься JSON и занесение этих данных в БД (регистрация)
def get_profile_vk(user_data):
    UserID = user_data[0]["id"]
    UserLastName = user_data[0]["last_name"]
    UserFirstName = user_data[0]["first_name"]
    UserSex = user_data[0]["sex"]
    UserCountry = user_data[0]["country"]['title']
    UserCity = user_data[0]["city"]['title']
    UserDomain = user_data[0]["domain"]
    UserPhoto200 = user_data[0]["photo_200"]
    register_new_user(UserID, UserLastName, UserFirstName, UserSex, UserCountry, UserCity, UserDomain, UserPhoto200)


# 2 - показать сколько осталось, 1 - название пары
def para(choose):
    x = choose
    h, m = dt.now().hour, dt.now().minute
    if ((h == 8 or h == 9) and m <= 30) and (x == 1):
        return 'Первая пара'
    elif ((h == 8 or h == 9) and m <= 30) and (x == 2):
        result = (9 * 60 + 30) - (h * 60 + m)
        return result
    elif ((h == 9 and m >= 40) or (h == 10) or (h == 11 and m <= 10)) and (x == 1):
        return 'Вторая пара'
    elif ((h == 9 and m >= 40) or (h == 10) or (h == 11 and m <= 10)) and (x == 2):
        result = (11 * 60 + 10) - (h * 60 + m)
        return result
    elif ((h == 11 and m >= 40) or (h == 12) or (h == 13 and m <= 10)) and (x == 1):
        return 'Третья пара'
    elif ((h == 11 and m >= 40) or (h == 12) or (h == 13 and m <= 10)) and (x == 2):
        result = (13 * 60 + 10) - (h * 60 + m)
        return result
    elif ((h == 13 and m >= 30) or (h == 14) or (h == 15 and m <= 0)) and (x == 1):
        return 'Четвертая пара'
    elif ((h == 13 and m >= 30) or (h == 14) or (h == 15 and m <= 0)) and (x == 2):
        result = (15 * 60 + 0) - (h * 60 + m)
        return result
    elif ((h == 15 and m >= 10) or (h == 16 and m <= 40)) and (x == 1):
        return 'Пятая пара'
    elif ((h == 15 and m >= 10) or (h == 16 and m <= 40)) and (x == 2):
        result = (16 * 60 + 40) - (h * 60 + m)
        return result
    elif ((h == 16 and m >= 50) or (h == 17) or (h == 18 and m <= 20)) and (x == 1):
        return 'Шестая пара'
    elif ((h == 16 and m >= 50) or (h == 17) or (h == 18 and m <= 20)) and (x == 2):
        result = (18 * 60 + 20) - (h * 60 + m)
        return result
    elif ((h == 18 and m >= 20) or (h == 19) or (h == 20 and m <= 0)) and (x == 1):
        return 'Седьмая пара'
    elif ((h == 18 and m >= 20) or (h == 19) or (h == 20 and m <= 0)) and (x == 2):
        result = (20 * 60 + 0) - (h * 60 + m)
    # else: result = 'Сейчач перемена!'
        return result



# Функция сообщения для кнопки НЕДЕЛЯ
def weekFull():
    wb1 = load_workbook('src/weeks.xlsx')
    sheet_ranges = wb1['sheet1']
    column_d = sheet_ranges['D']
    column_c = sheet_ranges['C']
    column_g = sheet_ranges['G']
    column_h = sheet_ranges['H']
    weeksTodayStr = datetime.datetime.now()
    weekTodayStr = weeksTodayStr.strftime("%U")
    weekTodayInt = int(weekTodayStr)
    weekTodayInt = weekTodayInt + 0
    for i in range(len(column_c)):
        if column_d[i].value == weekTodayInt:
            week = column_c[i].value
            break
        elif column_d[i].value is None:
            break
    for i in range(len(column_c)):
        if column_d[i].value == weekTodayInt:
            week1 = column_g[i].value
            break
        elif column_d[i].value is None:
            break
    for i in range(len(column_c)):
        if column_d[i].value == weekTodayInt:
            week2 = column_h[i].value
            break
        elif column_d[i].value is None:
            break
    return '📅 Сейчас идёт ' + str(week) + '-ая неделя' + '\nНачалась: ' + str(week1) + '\nЗакончится: ' + str(week2)
