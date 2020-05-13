#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from vk_api.longpoll import *

wb = openpyxl.load_workbook(filename='src/raspisanie.xlsx')


def parse_abitur_magistr():
    url = "http://www.pgsga.ru/abitur/bachelor/#abitur_spiski-postupayushchih"
    table = pd.read_html(url)[4]
    table.to_excel("src/abitur_magistr.xlsx")


def view_abitur_magistr():
    parse_abitur_magistr()
    wb1 = load_workbook('src/abitur_magistr.xlsx')
    sheet_ranges = wb1['Sheet1']
    column_b = sheet_ranges['B']
    column_c = sheet_ranges['C']
    column_d = sheet_ranges['D']
    column_e = sheet_ranges['E']
    column_f = sheet_ranges['F']
    column_g = sheet_ranges['G']

    magistrList = []
    for i in range(2, len(column_c)):
        if column_c[i].value is not None:
            # f = ' ' + column_d[i].value.replace(',', '\n')
            magistrList.append(
                '❗' + column_c[i].value + '\n✅ ' + column_d[i].value + '\n' + column_e[i].value + ' (' + column_f[
                    i].value + ')\n' + column_g[i].value + '\n\n')
        elif column_b[i].value == '':
            break
    return magistrList


def parse_abitur_bakalavr():
    url = "http://www.pgsga.ru/abitur/bachelor/#abitur_spiski-postupayushchih"
    table = pd.read_html(url)[0]
    table.to_excel("src/abitur_bakalavr.xlsx")


def view_abitur_bakalavr():
    parse_abitur_bakalavr()
    wb1 = load_workbook('src/abitur_bakalavr.xlsx')
    sheet_ranges = wb1['Sheet1']
    column_b = sheet_ranges['B']
    column_c = sheet_ranges['C']
    column_d = sheet_ranges['D']
    column_e = sheet_ranges['E']
    # column_f = sheet_ranges['F']
    # column_g = sheet_ranges['G']
    column_k = sheet_ranges['K']
    column_l = sheet_ranges['L']

    bakalavrList = []
    for i in range(5, len(column_c)):
        if column_c[i].value is not None:
            bakalavrList.append(
                '❗' + column_c[i].value + '\n✅ ' + column_d[i].value + '\n' + column_e[
                    i].value + '\nВсего мест ОЧНО: ' + column_k[i].value + '\nВсего мест ЗАОЧНО: ' + column_l[
                    i].value + '\n\n')
        elif column_b[i].value == '':
            break
    return bakalavrList


def get_mon(group):
    sheet = group
    ws = wb[sheet.lower()]
    column_mon = ws['A']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))


def get_tue(group):
    sheet = group
    ws = wb[sheet.lower()]
    column_mon = ws['B']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))


def get_wed(group):
    sheet = group
    ws = wb[sheet.lower()]
    column_mon = ws['C']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))


def get_thu(group):
    sheet = group
    ws = wb[sheet.lower()]
    column_mon = ws['D']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))


def get_fri(group):
    sheet = group
    ws = wb[sheet.lower()]
    column_mon = ws['E']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))


def get_sab(group):
    sheet = group
    ws = wb[sheet.lower()]
    column_mon = ws['F']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))
